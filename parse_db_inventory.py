#!/usr/bin/env python3
"""
Parse DB_LIST_MTN.xlsx and generate Ansible inventory for all database types.
Skips rows highlighted in red (decommissioned servers).
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import yaml
import os

EXCEL_FILE = 'DB_LIST_MTN.xlsx'
OUTPUT_DIR = 'ansible/inventory/mtn_databases'


def get_red_highlighted_rows(sheet):
    """Get row numbers that are highlighted in red (decommissioned)."""
    red_rows = set()
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Skip header
        for cell in row:
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                color = str(cell.fill.fgColor.rgb)
                # Check for red-ish colors (FF0000, FFFF0000, etc.)
                if color in ['FFFF0000', 'FF0000', 'FFFF6666', 'FFFF3333', 'FFCC0000']:
                    red_rows.add(row_idx)
                    break
    return red_rows


def parse_sheet(wb, sheet_name, df, db_type):
    """Parse a sheet and return hosts dict, filtering out red rows."""
    ws = wb[sheet_name]
    red_rows = get_red_highlighted_rows(ws)
    print(f"  - Found {len(red_rows)} decommissioned (red) rows in {sheet_name}")
    
    hosts = {'prod': {}, 'uat': {}}
    
    # Column mappings per db type
    col_map = {
        'mysql': {'host': 'HOST', 'ip': 'IP Address', 'port': 'PORT', 'env': 'Environment', 'role': 'Role', 'app': 'Application', 'version': 'DB version'},
        'mongo': {'host': 'hostname', 'ip': 'IP Address', 'port': 'PORT', 'env': 'Environment', 'role': 'type', 'app': 'Application ', 'version': 'version'},
        'oracle': {'host': 'Host', 'ip': 'IP', 'port': None, 'env': 'Type', 'role': None, 'app': 'DB', 'version': 'Version'},
        'mssql': {'host': 'Server Name', 'ip': 'IP', 'port': 'Port', 'env': 'Prod/UAT', 'role': None, 'app': 'Service', 'version': 'Version'},
        'postgres': {'host': 'HOST', 'ip': 'IP Address', 'port': 'PORT', 'env': 'Environment', 'role': 'Role', 'app': 'Application', 'version': 'DB version'},
        'cassandra': {'host': 'Server Name', 'ip': 'IP Adddress', 'port': 'CQL PORT', 'env': 'Environment', 'role': None, 'app': 'db name', 'version': 'db version'},
    }
    
    mapping = col_map.get(db_type, col_map['mysql'])
    
    for idx, row in df.iterrows():
        # Skip if in red rows (excel row = pandas idx + 2 because of 0-indexing and header)
        excel_row = idx + 2
        if excel_row in red_rows:
            continue
        
        # Get host and IP
        host_col = mapping['host']
        ip_col = mapping['ip']
        
        if host_col not in df.columns or ip_col not in df.columns:
            continue
            
        host = str(row.get(host_col, '')).strip()
        ip = str(row.get(ip_col, '')).strip()
        
        if not host or host == 'nan' or not ip or ip == 'nan':
            continue
        
        # Get other fields
        env = str(row.get(mapping.get('env', 'Environment'), '')).strip().upper()
        port_val = row.get(mapping.get('port'), None) if mapping.get('port') else None
        role = str(row.get(mapping.get('role'), '')) if mapping.get('role') else ''
        app = str(row.get(mapping.get('app'), '')) if mapping.get('app') else ''
        version = str(row.get(mapping.get('version'), '')) if mapping.get('version') else ''
        
        # Clean values
        if role in ['nan', 'None', '']: role = 'standalone'
        if app in ['nan', 'None', '']: app = ''
        if version in ['nan', 'None', '']: version = ''
        
        # Parse port
        try:
            port = int(float(port_val)) if port_val and str(port_val) not in ['nan', 'None', ''] else None
        except:
            port = None
        
        # Build host vars
        host_vars = {'ansible_host': ip}
        if port:
            host_vars[f'{db_type}_port'] = port
        if role != 'standalone':
            host_vars[f'{db_type}_role'] = role
        if app:
            host_vars[f'{db_type}_application'] = app
        if version:
            host_vars[f'{db_type}_version'] = version
        
        # Determine environment
        if 'PROD' in env or 'PRD' in env:
            hosts['prod'][host] = host_vars
        elif 'UAT' in env or 'DEV' in env or 'TEST' in env:
            hosts['uat'][host] = host_vars
        else:
            # Default to prod if environment unclear
            hosts['prod'][host] = host_vars
    
    return hosts


def generate_inventory(db_type, hosts):
    """Generate YAML inventory file for a database type."""
    inventory = {
        'all': {
            'children': {
                f'{db_type}_servers': {
                    'children': {
                        f'{db_type}_prod': {
                            'hosts': hosts['prod']
                        },
                        f'{db_type}_uat': {
                            'hosts': hosts['uat']
                        }
                    }
                }
            }
        }
    }
    return inventory


def main():
    print("Loading Excel file...")
    wb = load_workbook(EXCEL_FILE)
    xlsx = pd.ExcelFile(EXCEL_FILE)
    
    print(f"Available sheets: {wb.sheetnames}")
    
    # Create output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # Mapping of sheet names to db types
    sheet_mapping = {
        'MYSQL': 'mysql',
        'MONGO': 'mongo',
        'ORACLE': 'oracle',
        'SQL SERVER': 'mssql',
        'POSTGRES': 'postgres',
        'CASSANDRA': 'cassandra',
    }
    
    summary = {}
    
    for sheet_name, db_type in sheet_mapping.items():
        if sheet_name not in wb.sheetnames:
            print(f"\n[SKIP] Sheet '{sheet_name}' not found")
            continue
        
        print(f"\n[PROCESSING] {sheet_name} -> {db_type}")
        df = pd.read_excel(xlsx, sheet_name=sheet_name)
        print(f"  - Total rows: {len(df)}")
        
        hosts = parse_sheet(wb, sheet_name, df, db_type)
        
        prod_count = len(hosts['prod'])
        uat_count = len(hosts['uat'])
        total = prod_count + uat_count
        
        print(f"  - Active: {total} ({prod_count} prod, {uat_count} uat)")
        
        if total > 0:
            inventory = generate_inventory(db_type, hosts)
            output_file = os.path.join(OUTPUT_DIR, f'{db_type}_hosts.yml')
            
            with open(output_file, 'w') as f:
                yaml.dump(inventory, f, default_flow_style=False, sort_keys=False, allow_unicode=True)
            
            print(f"  - Saved to: {output_file}")
            summary[db_type] = {'prod': prod_count, 'uat': uat_count}
    
    # Print summary
    print("\n" + "="*50)
    print("SUMMARY")
    print("="*50)
    for db_type, counts in summary.items():
        print(f"  {db_type.upper():12} -> {counts['prod']:4} prod, {counts['uat']:4} uat")
    
    total_prod = sum(c['prod'] for c in summary.values())
    total_uat = sum(c['uat'] for c in summary.values())
    print(f"  {'TOTAL':12} -> {total_prod:4} prod, {total_uat:4} uat")


if __name__ == '__main__':
    main()
